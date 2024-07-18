import os
import time
import zipfile
import json
import streamlit as st
from tqdm import tqdm
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from PyPDF2 import PdfReader, PdfWriter
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import shutil
import glob


# Page setup
st.set_page_config(layout='wide')


# Azure Form Recognizer configuration
azure_endpoint = st.secrets["azure"]["az_endpoint"]
azure_key  = st.secrets["azure"]["az_key"]
model_id = st.secrets["azure"]["az_model_id"]

def download_chromedriver():
    os.system('sbase install chromedriver')
    os.system('ln -s /home/appuser/.local/lib/python3.7/site-packages/seleniumbase/drivers/chromedriver /usr/bin/chromedriver')

@st.cache_resource
def init_driver():
    download_chromedriver()

    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--remote-debugging-port=9222')
    
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def download_zip(folder_url):
    driver = init_driver()
    # Get the current list of files in the Downloads folder
    downloads_folder = os.path.expanduser('~/Downloads')

    # Create the downloads folder if it doesn't exist
    if not os.path.exists(downloads_folder):
        os.makedirs(downloads_folder)
    before_download = set(os.listdir(downloads_folder))
    
    driver.get(folder_url)
    time.sleep(5)
    
    # Locate and click the download button
    download_button = driver.find_element(By.XPATH, '//*[@id="appRoot"]/div[1]/div[2]/div/div[2]/div[2]/div[2]/div[1]/div/div/div/div/div/div/div[1]/div[1]/button/span')
    download_button.click()
    
    time.sleep(5)  # Wait a bit for the download to start

    # Wait for the new ZIP file to appear in the Downloads folder
    zip_name = None
    for _ in range(30):  # Wait for up to 30 seconds
        time.sleep(1)
        # List files in the Downloads folder
        after_download = set(os.listdir(downloads_folder))
        new_files = after_download - before_download
        zip_files = [f for f in new_files if f.lower().endswith(".zip")]
        if zip_files:
            zip_name = max(zip_files, key=lambda x: os.path.getmtime(os.path.join(downloads_folder, x)))
            break

    if zip_name is None:
        raise FileNotFoundError("ZIP file was not downloaded.")

    return os.path.join(downloads_folder, zip_name)  # Return the full path to the ZIP file
    

def extract_zip(zip_path, extract_to):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)

def find_pdf_files(folder):
    pdf_files = []
    for root, _, files in os.walk(folder):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdf_files.append(os.path.join(root, file))
    return pdf_files
    
    
def split_pdf_in_pairs(pdf_path):
    try:
        reader = PdfReader(pdf_path)
        num_pages = min(len(reader.pages), 11)  # Only process up to page 11
        pairs = []
        for i in tqdm(range(0, num_pages, 2), desc="Splitting PDF into pairs"):
            writer = PdfWriter()
            writer.add_page(reader.pages[i])
            if i + 1 < num_pages:
                writer.add_page(reader.pages[i + 1])
            pair_path = f"temp_pair_{i // 2}.pdf"
            with open(pair_path, "wb") as f:
                writer.write(f)
            pairs.append(pair_path)
        return pairs
    except Exception as e:
        print(f"Error reading PDF file: {e}")
        return []

def analyze_pdf_pair(document_analysis_client, model_id, pdf_path):
    with open(pdf_path, "rb") as f:
        poller = document_analysis_client.begin_analyze_document(model_id=model_id, document=f)
    result = poller.result()
    data = []
    for idx, document in enumerate(result.documents):
        fields_with_values = {
            name: (field.value if field.value else field.content) 
            for name, field in document.fields.items() 
            if (field.value or field.content) and field.value != "unselected"
        }
        doc_data = {
            "document_number": idx + 1,
            "doc_type": document.doc_type,
            "confidence": document.confidence,
            "model_id": result.model_id,
            "fields": fields_with_values
        }
        data.append(doc_data)
    return data

def save_data(data, json_path):
    with open(json_path, "w", encoding='utf-8', errors='ignore') as json_file:
        json.dump(data, json_file, indent=4)
        
        
        

def split_text(text):
    """
    Splits the text by clauses.
    """
    pattern = r'(D\d+\.\d+(?:[-\/\s&)]*D\d+\.\d+)*)'
    text = text.encode('ascii', errors='ignore').decode('utf-8')
    sections = re.split(pattern, text)
    result = {}
    other_comments = []

    i = 0
    while i < len(sections):
        if i % 2 == 0:  # even index, this is the text part
            if sections[i].strip():
                other_comments.append(sections[i].strip())
        else:  # odd index, this is the key part
            key = sections[i].strip() if sections[i] else None
            value = sections[i + 1].strip() if (i + 1) < len(sections) and sections[i + 1] else ""
            if key:
                result[key] = value
                i += 1  # skip the value part already processed
                # Remove this value from other_comments if it was added before
                if other_comments and other_comments[-1] == value:
                    other_comments.pop()
        i += 1

    if other_comments:
        result["Other"] = " ".join(other_comments)
    
    return result

def clean_text(text):
    """
    Removes non-ASCII characters from the text.
    """
    return re.sub(r'[^\x00-\x7F]+', '', text) if isinstance(text, str) else text

def process_json_file(file_path, output_directory):
    # Extract the characters before the first underscore in the original file path name
    file_prefix = os.path.basename(file_path).split("_", 1)[0]

    # Load the original JSON file
    with open(file_path) as f:
        d = json.load(f)

    # Create a copy of the original data
    updated_data = d.copy()

    # Iterate over the list and process the "15 precisions" field
    for doc in updated_data:
        fields = doc.get('fields', {})
        if '15 precisions' in fields:
            text = fields['15 precisions']
            split_result = split_text(text)
            
            # Create a new JSON object with the split results
            comments = {}
            for key, value in split_result.items():
                comments[key + " Comments"] = value
            
            # Add the new JSON object to the existing JSON structure
            if 'Comments' in fields:
                fields['Comments'].update(comments)
            else:
                fields['Comments'] = comments
            
           
    
            # Update the original JSON data with the new results
            doc['fields'] = fields
    
    # Create a new list to hold the updated data
    cleaned_data = []
    
    # Iterate over each document in the original data
    for doc in d:
        # Extract only the "fields" from each document
        fields = doc.get('fields', {})
        # Clean the keys and values of the fields dictionary
        cleaned_fields = {clean_text(k): clean_text(v) for k, v in fields.items()}
    
        cleaned_doc = {'fields': cleaned_fields}
        
        # Append the updated document to the list
        cleaned_data.append(cleaned_doc)
    
    # Flattening and combining the data into one dictionary
    combined_data = {}
    
    for item in cleaned_data:
        combined_data.update(item["fields"])
    
    # Ensure that "DV Number" is used as the ID field and the partition key
    if "DV Number" in combined_data:
        combined_data["id"] = combined_data.pop("DV Number")
        combined_data["DVNumber"] = combined_data["id"]
    
    # Write the updated data to a new JSON file
    output_file_path = os.path.join(output_directory, file_prefix + "_updated.json")
    with open(output_file_path, 'w') as f:
        json.dump(combined_data, f, indent=4)
    
    print(f"Updated JSON written to {output_file_path}.")
    
    return combined_data

def json_to_dataframe(file_path):
    with open(file_path, 'r') as f:
        data = json.load(f)
        # Convert JSON to a flat DataFrame
        df = pd.json_normalize(data)
    return df

def merge_json_files_to_csv(input_directory, output_file_path):
    all_dataframes = []
    
    # Read each JSON file and convert it to a DataFrame
    for file_name in os.listdir(input_directory):
        if file_name.endswith('.json'):
            file_path = os.path.join(input_directory, file_name)
            df = json_to_dataframe(file_path)
            all_dataframes.append(df)
    
    # Merge all DataFrames into a single DataFrame
    merged_df = pd.concat(all_dataframes, ignore_index=True)

    # Sort columns alphabetically
    merged_df = merged_df.reindex(sorted(merged_df.columns), axis=1)
    
    # Save the merged DataFrame to a CSV file
    merged_df.to_csv(output_file_path, index=False)


def create_excel_report(csv_file_path, excel_output_path):
    df = pd.read_csv(csv_file_path)

    # Define the section mapping
    section_mapping = {
        '2': 'General Information',
        '3': 'Land (Soil)',
        '4': 'Damage caused by water',
        '5': 'Basement & foundation',
        '6': 'Undesirable animals',
        '7': 'Interior air quality',
        '8': 'Roof',
        '9': 'Plumbing and drainage',
        '10': 'Energy',
        '11': 'Telecommunications',
        '12': 'Heating, air conditioning & ventilation',
        '13': 'Inspection & other expert reports',
        '14': 'Other information',
        '15': 'Details'
    }
    
    # Summary Page
    # Initialize the summary DataFrame
    summary_data = {
        'Address': [],
        'Year of construction': [],
        'Year of acquisition': [],
        'General Information': [],
        'Land (Soil)': [],
        'Damage caused by water': [],
        'Basement & foundation': [],
        'Undesirable animals': [],
        'Interior air quality': [],
        'Roof': [],
        'Plumbing and drainage': [],
        'Energy': [],
        'Telecommunications': [],
        'Heating, air conditioning & ventilation': [],
        'Inspection & other expert reports': [],
        'Other information': []
    }
    
    # Initialize an empty dictionary to store section counts for each address
    address_section_counts = {}
    
    # Process each property
    for idx, row in df.iterrows():
        address = row['Adresse de lmmeuble']
        year_construction = row['2.2 annee de construction']
        year_acquisition = row['2.1 annee acquis']
    
        # Initialize counts for each section for this address
        section_counts = {section: 0 for section in section_mapping.values() if section != 'Details'}
    
        # Count the number of selected questions for each section
        for col in df.columns:
            if col.endswith('O') and 'selected' in str(row[col]):  # Check if column ends with 'O' and is selected
                for prefix, section in section_mapping.items():
                    if col.startswith(prefix) and section != 'Details':  # Exclude the 'Details' section
                        section_counts[section] += 1
                        break
    
        # Store the section counts for this address
        address_section_counts[address] = section_counts
    
        # Append the results to the summary data
        summary_data['Address'].append(address)
        summary_data['Year of construction'].append(year_construction)
        summary_data['Year of acquisition'].append(year_acquisition)
        for section, count in section_counts.items():
            summary_data[section].append(count)
    
    # Ensure all lists in summary_data have the same length
    max_length = max(len(v) for v in summary_data.values())
    for k, v in summary_data.items():
        if len(v) < max_length:
            v.extend([None] * (max_length - len(v)))
    
    # Create the summary DataFrame
    summary_df = pd.DataFrame(summary_data)
    
    # Set the Address column as the index
    summary_df.set_index('Address', inplace=True)
    
    # Transpose the DataFrame
    transposed_summary_df = summary_df.transpose()
    
    # Rename the index to 'Address'
    transposed_summary_df.index.name = 'Address'
    
    # Reset the index to make the transposed DataFrame more readable
    transposed_summary_df.reset_index(inplace=True)
    
    # Define the file path where you want to save the Excel file
    output_file_path = 'processed_files/RE_ComparisonReport.xlsx'
    
    # Prepare the Details DataFrame
    # Assuming the address column is named 'Address'
    address_column = 'Adresse de lmmeuble'
    
    # Set the Address column as the headers
    addresses = df[address_column].tolist()
    
    # Drop the Address column from the original DataFrame
    df = df.drop(columns=[address_column])
    
    # Transpose the DataFrame
    df_transposed = df.transpose()
    
    # Reset index to make the first column (questions) a regular column
    df_transposed.reset_index(inplace=True)
    
    # Rename columns appropriately
    df_transposed.columns = ['Question'] + addresses
    
    # Extract section and question number from the 'Question' column
    def extract_section(question):
        if question.startswith('15 precisions') or 'Comments' in question:
            return '15'
        parts = question.split('.')
        if len(parts) > 1:
            return parts[0]
        else:
            return ''
    
    def extract_question_number(question):
        try:
            if 'Comments' in question:
                return 99999  # Ensure 'Comments' are sorted last
            parts = question.split('.')
            if len(parts) > 1:
                num_parts = parts[1].split()
                return float(num_parts[0])
            else:
                return float('inf')  # Use 'inf' to ensure unmapped questions are sorted last within their section
        except ValueError:
            return float('inf')  # Return 'inf' if conversion fails
    
    df_transposed['Section'] = df_transposed['Question'].apply(extract_section)
    df_transposed['QuestionNumber'] = df_transposed['Question'].apply(extract_question_number)
    
    # Map section numbers back to section names, defaulting to 'Other Information' for unmapped sections
    df_transposed['SectionName'] = df_transposed['Section'].map(section_mapping).fillna('Other Information')
    
    # Filter out rows that don't have a valid section mapping
    df_transposed = df_transposed[df_transposed['Section'].isin(section_mapping.keys())]
    
    # Sort by section name and then by question number within each section
    def sort_questions(row):
        section = row['Section']
        question_number = row['QuestionNumber']
        if section == '15':
            if 'Comments' in row['Question']:
                return (15, 99999)  # Sort comments last within '15' section
            elif row['Question'].startswith('15.1') or row['Question'].startswith('15 precisions'):
                return (15, 0)  # Sort '15.1' and '15 precisions' first within '15' section
            else:
                return (15, question_number)  # Sort other '15' questions by question number
        return (int(section), question_number)
    
    df_transposed['SortKey'] = df_transposed.apply(sort_questions, axis=1)
    df_transposed.sort_values(by=['SortKey'], inplace=True)
    
    # Insert section headers
    df_list = []
    current_section = None
    num_columns = len(df_transposed.columns)
    for _, row in df_transposed.iterrows():
        section_name = row['SectionName']
        if section_name != current_section:
            df_list.append(pd.DataFrame([[section_name] + [''] * (num_columns - 1)], columns=df_transposed.columns))
            current_section = section_name
        df_list.append(pd.DataFrame([row.values], columns=df_transposed.columns))
    
    df_with_sections = pd.concat(df_list, ignore_index=True)
    
    # Drop the 'Section', 'SectionName', 'QuestionNumber', and 'SortKey' columns as they are no longer needed in the output
    df_with_sections.drop(columns=['Section', 'SectionName', 'QuestionNumber', 'SortKey'], inplace=True)
    
    # Replace 'selected' with 'YES' across all columns
    df_with_sections.replace('selected', 'YES', inplace=True)
    
    # Save both DataFrames to different sheets in the same Excel file
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        transposed_summary_df.to_excel(writer, sheet_name='Summary', index=False)
        df_with_sections.to_excel(writer, sheet_name='Details', index=False)
    
    # Load the workbook to apply formatting
    wb = load_workbook(output_file_path)
    
    # Summary Sheet Formatting
    ws_summary = wb['Summary']
    ws_summary.insert_rows(1)
    ws_summary['A1'] = 'Summary Report'
    title_font = Font(size=24, bold=True)
    ws_summary['A1'].font = title_font
    
    # Formatting address headers in the summary sheet
    address_font = Font(size=12, bold=True, color="FFFFFF")
    address_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")  # Dark blue background
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(2, ws_summary.max_column + 1):
        cell = ws_summary.cell(row=2, column=col)  # Address headers start from the second row
        cell.font = address_font
        cell.fill = address_fill
        cell.alignment = center_alignment
    
    # Right alignment for all cells after the address headers
    for col in range(2, ws_summary.max_column + 1):
        for row in range(3, ws_summary.max_row + 1):  # Start from row 3 to skip title and headers
            ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal="right")
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    for col in range(2, ws_summary.max_column + 1):  # Start from 2 to skip the 'Category' column
        for row in range(6, ws_summary.max_row + 1):  # Start from row 6 to skip headers and the title
            cell = ws_summary.cell(row=row, column=col)
            try:
                value = int(cell.value)
                if value <= 3:
                    cell.fill = green_fill
                elif 3 < value <= 6:
                    cell.fill = yellow_fill
                elif value > 6:
                    cell.fill = red_fill
            except (ValueError, TypeError):
                pass
    
    fixed_width = 32  # Adjust this value as needed
    for column in ws_summary.columns:
        column_letter = column[0].column_letter  # Get the column letter
        ws_summary.column_dimensions[column_letter].width = fixed_width
    
    # Details Sheet Formatting
    ws_details = wb['Details']
    ws_details.insert_rows(1)
    ws_details['A1'] = 'Details Report'
    title_font = Font(size=24, bold=True)
    ws_details['A1'].font = title_font
    bold_font = Font(bold=True, size=14)
    blue_fill = PatternFill(start_color="B0E0E6", end_color="B0E0E6", fill_type="solid")  # Light blue fill
    
    # Apply light blue fill to the entire row of section headers
    for row in ws_details.iter_rows(min_row=2, max_row=ws_details.max_row):
        first_cell = row[0]
        if first_cell.value in section_mapping.values() or first_cell.value == 'Other Information':
            for cell in row:
                cell.fill = blue_fill
                cell.font = bold_font
    
    # Right alignment for all cells after the address headers
    for col in range(2, ws_details.max_column + 1):
        for row in range(3, ws_details.max_row + 1):  # Start from row 3 to skip title and headers
            ws_details.cell(row=row, column=col).alignment = Alignment(horizontal="right")
    
    fixed_width = 32  # Adjust this value as needed
    for column in ws_details.columns:
        column_letter = column[0].column_letter  # Get the column letter
        ws_details.column_dimensions[column_letter].width = fixed_width
    
    small_font = Font(size=8)
    for row in ws_details.iter_rows(min_row=2, max_row=ws_details.max_row):
        cell = row[0]
        if '15 precisions' in cell.value or 'Comments' in cell.value:
            for cell in row:
                cell.font = small_font
                cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping
            for cell in row:
                ws_details.column_dimensions[cell.column_letter].width = 32  # Set a larger column width
            ws_details.row_dimensions[cell.row].height = 100  # Adjust this value as needed for larger row height
    
    # Formatting address headers in the details sheet
    for col in range(2, ws_details.max_column + 1):
        cell = ws_details.cell(row=2, column=col)  # Address headers are in the first row
        cell.font = address_font
        cell.fill = address_fill
        cell.alignment = center_alignment
    
    # Save the formatted Excel file
    wb.save(output_file_path)
    
    print(f"Summary and Details report saved to {output_file_path}")
    # Completion message
    st.success(f"Analysis Complete, Comparison Report Generated in C:/Users/ {output_file_path}!")
    
def remove_json_files(folder_path):
    """
    Removes all JSON files in the specified folder.

    Args:
    folder_path (str): The path to the folder where JSON files should be removed.
    """
    # Get the list of all JSON files in the folder
    json_files = glob.glob(os.path.join(folder_path, "*.json"))
    
    # Loop through the list and remove each file
    for json_file in json_files:
        os.remove(json_file)
        print(f"Removed: {json_file}")    


# Define the local directory containing original JSON files
input_directory = "processed_files"
output_directory = "processed_files/cleanpdfs"
os.makedirs(output_directory, exist_ok=True)
# Define the input directory containing PROCESSED JSON files and the output CSV file path
input_directory_processed = "processed_files/cleanpdfs"
combined_csv_path = "processed_files/cleanpdfs/combinedJSONoutput.csv"
folder_path = "path/to/your/folder"

       

def main():
    custom_css = """
    <style>
        /* Set the background color to dark blue for the app */
        .stApp {
            background-color: #001f3f;
        }
        /* Customize the title */
        h1 {
            color: white !important; 
            font-family: 'Arial', sans-serif; 
            font-size: 5em; 
            text-shadow: 1px 1px 0 #FF00FF, -1px -1px 0 #FF00FF, 1px -1px 0 #FF00FF, -1px 1px 0 #FF00FF;
        }
        /* Customize the button color and border */
        .stButton button {
            background-color: white; 
            color: black; /* Green text */
            border: 2px solid #32CD32; /* Green border */
            padding: 10px 24px;
            font-size: 16px;
            border-radius: 8px;
        }
        /* Change button color on hover */
        .stButton button:hover {
            background-color: #32CD32; /* Green background on hover */
            color: white; /* White text on hover */
            border: 2px solid white; /* White border on hover */
        }
        /* Customize the input text box */
        .stTextInput input {
            background-color: #1E1E1E; /* Dark background */
            color: #FFFFFF; /* White text */
            border: 1px solid #FFFFFF; /* White border */
            padding: 10px;
            border-radius: 4px;
        }
        /* Customize input label */
        label.css-1cpxqw2 {
            color: #FFFFFF !important;
            font-family: Arial, sans-serif; 
            font-size: 3em !important; /* Larger font size */
        }
    </style>
    """

    # Embed the custom CSS in your Streamlit app
    st.markdown(custom_css, unsafe_allow_html=True)
    st.title("Real Estate Date Extract")
    
     # Sidebar for About section
    with st.sidebar:
        st.header("About this web app")
        st.write("""
        This application allows users to select files from SharePoint,
        analyze Real Estate PDF documents and generate a Comparison Report with the data in the RE forms. 
          
        """)
    
    # Input fields for SharePoint credentials and site details
    folder_url = st.text_input(":green[SharePoint Folder URL]")

    if st.button("Generate Comparison Report"):
        if folder_url:
            # Download the ZIP file from SharePoint
            zip_path = download_zip(folder_url)

            # Extract the ZIP file
            local_folder_path = "extracted_pdfs"  # Folder to extract PDFs
            os.makedirs(local_folder_path, exist_ok=True)
            
            extract_zip(zip_path, local_folder_path)

            # Close the driver
            driver.quit()
            
            document_analysis_client = DocumentAnalysisClient(endpoint=azure_endpoint, credential=AzureKeyCredential(azure_key))
            
            processed_folder_path = "processed_files"  # Folder for processed files
            os.makedirs(processed_folder_path, exist_ok=True)

            # Find all PDF files in the extracted folder
            pdf_files = find_pdf_files(local_folder_path)
            total_files = len(pdf_files)

            progress_bar = st.progress(0)
            status_text = st.empty()

            for idx, pdf_path in enumerate(pdf_files):
                file_name = os.path.basename(pdf_path)
                try:
                    if not os.path.isfile(pdf_path) or os.path.getsize(pdf_path) == 0:
                        st.warning(f"File {file_name} is corrupted or empty, skipping.")
                        continue

                    pdf_pairs = split_pdf_in_pairs(pdf_path)
                    if not pdf_pairs:
                        st.warning(f"Skipping {file_name} due to PDF read error.")
                        continue

                    all_data = []
                    pair_progress_bar = st.progress(0)  # Progress bar for pairs
                    pair_status_text = st.empty()  # Status text for pairs
                    for  pair_idx, pair_path in enumerate(pdf_pairs):
                        data = analyze_pdf_pair(document_analysis_client, model_id, pair_path)
                        all_data.extend(data)
                        os.remove(pair_path)
                        
                        # Custom progress bar for pairs
                        pair_progress = (pair_idx + 1) / len(pdf_pairs)
                        pair_progress_bar.markdown(
                            f"""
                            <div style="width: 100%; background-color: #ffffff; border-radius: 5px;">
                                <div style="width: {pair_progress * 100}%; background-color: #8D89F6; height: 20px; border-radius: 5px;"></div>
                            </div>
                            """, unsafe_allow_html=True
                        )

                        # Update status text for pairs
                        pair_status_text.text(f"Analyzing pair {pair_idx + 1} of {len(pdf_pairs)} ({pair_progress * 100:.2f}%)")

                    json_file_name = f"{os.path.splitext(file_name)[0]}_data.json"
                    json_path = os.path.join(processed_folder_path, json_file_name)
                    save_data(all_data, json_path)

                    os.rename(pdf_path, os.path.join(processed_folder_path, file_name))
                    
                except Exception as e:
                    st.error(f"An error occurred with file {file_name}: {e}")

                # Update progress bar and status text
                progress = (idx + 1) / total_files
                progress_bar.markdown(
                            f"""
                            <div style="width: 100%; background-color: #ffffff; border-radius: 5px;">
                                <div style="width: {progress * 100}%; background-color: #060270; height: 20px; border-radius: 5px;"></div>
                            </div>
                            """, unsafe_allow_html=True
                        )
                progress_bar.progress(progress)
                status_text.text(f"OVERALL PROCESSING - Processing file {idx + 1} of {total_files} ({progress * 100:.2f}%)")

            
            
            # Process each JSON file in the local directory
            for file_name in os.listdir(input_directory):
                if file_name.endswith('.json'):
                    file_path = os.path.join(input_directory, file_name)
                    combined_data = process_json_file(file_path, output_directory)


                    
            # Merge JSON files and generate the CSV file
            merge_json_files_to_csv(input_directory_processed, combined_csv_path)
            print(f"Combined CSV created in {combined_csv_path}.")   

            create_excel_report(combined_csv_path, output_directory)
            shutil.rmtree("extracted_pdfs")
            shutil.rmtree("processed_files/cleanpdfs")
            remove_json_files("processed_files/")
            print("code execution complete")
            
            

if __name__ == "__main__":
    main()
